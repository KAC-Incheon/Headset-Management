# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'headset6.ui'
#
# Created by: PyQt5 UI code generator 5.15.6
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

# from asyncio.windows_events import NULL
# from hashlib import new
# from operator import truediv
# from pickle import FALSE, TRUE
# from re import sub

###디자인 수정 중

import sys  # 모듈 불러오기
import os
import datetime
# from tkinter import Button
# from xmlrpc.client import boolean

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem, QDialog, QMessageBox, QComboBox, QFileDialog
from PyQt5 import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

import pymysql
import openpyxl

from DBConnect import DB
from openpyxl import load_workbook

from UIRegister import Ui_Resigster
from UIUpdate import Ui_Update


# conn, cur = None, None  # sql 연결 커서

class Ui_MainWindow(QWidget):

    def __init__(self):
        super().__init__()
        self.db = DB()
        self.btnStyleSet = "QPushButton{border-style:solid;border-width: 2px;border-color: #9DCFFF;border-radius: 3px;}QPushButton::hover{border-style:solid; border-width: 1px; color:white; background-color: #9DCFFF;border-radius: 3px;}"

    def setupUi(self, MainWindow):

        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(1036, 894)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(40, 10, 951, 311))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label = QtWidgets.QLabel(self.verticalLayoutWidget)
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setPointSize(13)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.tableWidget = QtWidgets.QTableWidget(self.verticalLayoutWidget)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        self.verticalLayout.addWidget(self.tableWidget)
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        self.tableWidget.setFont(font)
        self.groupBox = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox.setGeometry(QtCore.QRect(40, 340, 381, 280))

        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        self.groupBox.setFont(font)
        self.groupBox.setObjectName("groupBox")
        self.label_2 = QtWidgets.QLabel(self.groupBox)
        self.label_2.setGeometry(QtCore.QRect(130, 50, 201, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")

        self.textBrowser = QtWidgets.QTextBrowser(self.groupBox)
        self.textBrowser.setGeometry(QtCore.QRect(200, 50, 50, 30))
        self.textBrowser.setObjectName("textBrowser")

        self.textBrowser_2 = QtWidgets.QTextBrowser(self.groupBox)
        self.textBrowser_2.setGeometry(QtCore.QRect(200, 116, 50, 30))
        self.textBrowser_2.setObjectName("textBrowser_2")

        self.textBrowser_3 = QtWidgets.QTextBrowser(self.groupBox)
        self.textBrowser_3.setGeometry(QtCore.QRect(200, 180, 50, 30))
        self.textBrowser_3.setObjectName("textBrowser_3")

        self.label_4 = QtWidgets.QLabel(self.groupBox)
        self.label_4.setGeometry(QtCore.QRect(116, 110, 211, 41))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.groupBox)
        self.label_5.setGeometry(QtCore.QRect(140, 180, 131, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")

        self.groupBox_2 = QtWidgets.QGroupBox(self.centralwidget)
        self.groupBox_2.setGeometry(QtCore.QRect(450, 330, 430, 530))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        self.groupBox_2.setFont(font)
        self.groupBox_2.setObjectName("groupBox_2")
        self.textBrowser_6 = QtWidgets.QTextBrowser(self.groupBox_2)
        self.textBrowser_6.setGeometry(QtCore.QRect(680, 79, 121, 31))
        self.textBrowser_6.setObjectName("textBrowser_6")
        self.textEdit = QtWidgets.QTextEdit(self.groupBox_2)
        self.textEdit.setGeometry(QtCore.QRect(220, 180, 100, 30))
        self.textEdit.setObjectName("textEdit")
        self.label_6 = QtWidgets.QLabel(self.groupBox_2)
        self.label_6.setGeometry(QtCore.QRect(84, 180, 71, 31))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(False)
        font.setWeight(50)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.groupBox_2)
        self.label_7.setGeometry(QtCore.QRect(100, 40, 71, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.comboBox = QtWidgets.QComboBox(self.groupBox_2)
        self.comboBox.setGeometry(QtCore.QRect(220, 40, 100, 30))
        self.comboBox.setObjectName("comboBox")
        self.comboBox_2 = QtWidgets.QComboBox(self.groupBox_2)
        self.comboBox_2.setGeometry(QtCore.QRect(220, 110, 100, 30))
        self.comboBox_2.setObjectName("comboBox")
        self.label_8 = QtWidgets.QLabel(self.groupBox_2)
        self.label_8.setGeometry(QtCore.QRect(100, 320, 71, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_8.setFont(font)
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.groupBox_2)
        self.label_9.setGeometry(QtCore.QRect(84, 110, 81, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_9.setFont(font)
        self.label_9.setObjectName("label_9")
        self.label_10 = QtWidgets.QLabel(self.groupBox_2)
        self.label_10.setGeometry(QtCore.QRect(114, 250, 81, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_10.setFont(font)
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self.groupBox_2)
        self.label_11.setGeometry(QtCore.QRect(100, 390, 71, 31))
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        self.label_11.setFont(font)
        self.label_11.setObjectName("label_11")
        # self.textEdit_3 = QtWidgets.QTextEdit(self.groupBox_2)
        # self.textEdit_3.setGeometry(QtCore.QRect(220, 40, 151, 41))
        # self.textEdit_3.setObjectName("textEdit_3")
        self.textEdit_4 = QtWidgets.QTextEdit(self.groupBox_2)
        self.textEdit_4.setGeometry(QtCore.QRect(220, 320, 100, 30))
        self.textEdit_4.setObjectName("textEdit_4")
        # self.textEdit_5 = QtWidgets.QTextEdit(self.groupBox_2)
        # self.textEdit_5.setGeometry(QtCore.QRect(220, 110, 151, 41))
        # self.textEdit_5.setObjectName("textEdit_5")
        self.textEdit_6 = QtWidgets.QTextEdit(self.groupBox_2)
        self.textEdit_6.setGeometry(QtCore.QRect(220, 250, 100, 30))
        self.textEdit_6.setObjectName("textEdit_6")
        self.textEdit_7 = QtWidgets.QTextEdit(self.groupBox_2)
        self.textEdit_7.setGeometry(QtCore.QRect(220, 390, 100, 30))
        self.textEdit_7.setObjectName("textEdit_7")
        self.pushButton = QtWidgets.QPushButton(self.groupBox_2)
        self.pushButton.setGeometry(QtCore.QRect(295, 460, 101, 41))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(910, 370, 101, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_3 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(910, 430, 101, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_4 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(910, 490, 101, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_5 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_5.setGeometry(QtCore.QRect(910, 790, 101, 41))
        font = QtGui.QFont()
        font.setFamily("맑은 고딕")
        font.setBold(True)
        font.setWeight(75)
        font.setPointSize(8)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setObjectName("pushButton_5")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1036, 18))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.pushButton.clicked.connect(MainWindow.showlist)  ### type: ignore 함수연결 ###
        self.pushButton_3.clicked.connect(MainWindow.delete)
        self.pushButton_2.clicked.connect(self.dialog_open)  # 수정창 띄우기
        self.pushButton_4.clicked.connect(self.register_open)  # 등록창 띄우기
        self.pushButton_5.clicked.connect(self.export)  # 엑셀만들기
        # 검색버튼
        self.pushButton.setStyleSheet(self.btnStyleSet)
        # 수정버튼
        self.pushButton_2.setStyleSheet(self.btnStyleSet)
        # 삭제버튼
        self.pushButton_3.setStyleSheet(self.btnStyleSet)
        # 등록버튼
        self.pushButton_4.setStyleSheet(self.btnStyleSet)
        self.pushButton_5.setStyleSheet(self.btnStyleSet)

        self.tableWidget.verticalHeader().sectionClicked.connect(self.row_click)  # 행 헤더 클릭

        # 행 번호 말고 다른곳 클릭하면 선택한 값 위치 리셋
        self.tableWidget.cellClicked.connect(self.cellclicked_event)
        self.tableWidget.horizontalHeader().sectionClicked.connect(self.headerclicked_event)
        self.tableWidget.setCornerButtonEnabled(False)  # 왼쪽 상단 코너버튼 클릭 끄기
        self.tableWidget.verticalHeader().setFixedWidth(46)  ##열 헤더 너비 고정

        self.tableWidget.setStyleSheet("border-style:solid;"
                                       "border-width: 2px;"
                                       "border-color: #9DCFFF;"
                                       "border-radius: 3p")

        self.tableWidget.setEditTriggers(QAbstractItemView.NoEditTriggers)  ###테이블 안에서 수정 금지
        # self.tableWidget.setSelectionMode(QAbstractItemView.NoSelection)   ###데이터 선택 금지

        self.tableWidget.setSortingEnabled(True)  # 열 정렬

        # 지급수량 세기
        self.textBrowser.setText(
            str(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state = '지급';")[0]))  ##### 개수 표출

        ##### 미지급수량 세기
        self.textBrowser_2.setText(
            str(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('반납','none');")[0]))  ##### 개수 표출

        ##### 총수량 세기
        self.textBrowser_3.setText(str(len(self.db.execute_all("select * from v_new;"))))  ##### 개수 표출

        ###콤보박스 리스트 추가
        self.comboBox.addItem('')

        for model in self.db.execute_all(
                "SELECT DISTINCT model_name FROM v_now WHERE NOT model_name = 'none' ORDER BY model_name DESC ;"):
            self.comboBox.addItem(model[0])

        self.comboBox.addItem('none')
        self.comboBox.currentIndexChanged.connect(self.combo_change)

        cur = self.db.conn.cursor()

        ####지급여부 콤보박스2 리스트 설정
        self.comboBox_2.addItem('')
        for given in self.db.execute_all("SELECT DISTINCT state FROM v_now ORDER BY state DESC;"):
            self.comboBox_2.addItem(given[0])

        self.comboBox_2.currentIndexChanged.connect(self.combo2_change)  ###콤보박스 선택 시 함수

    def combo_change(self):  ###모델명

        combo = self.comboBox.currentText()
        return combo

    def combo2_change(self):  ###### 지급or반납
        combo = self.comboBox_2.currentText()

        return combo

    def register_open(self):  # 등록 창 오픈

        self.close()  ###메인윈도우 닫음

        self.register = Ui_Resigster()  ## 수정창 클래스 생성
        # except :  QMessageBox.warning(self,'알림','수정할 데이터를 선택하세요')

        self.register.exec()  ###창 지속
        self.show()  ###  창 닫으면 메인윈도우 다시 표출

        ########################    등록 후 다시 검색 ####################################
        ##### 지급수량 세기

        self.textBrowser.setText(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state = '지급';")[0])  ##### 개수 표출
        ##### 미지급수량 세기 

        self.textBrowser_2.setText(
            self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('반납','none');")[0])  ##### 개수 표출
        ##### 총수량 세기
        self.textBrowser_3.setText(len(self.db.execute_one("select * from v_new")))  ##### 개수 표출

        self.tableWidget.clear()

        # if not any(self.getConditions()):  ########## 모든 칸에 입력 없을 경우 전체 헤드셋 표출
        #     self.showTableListAll("select * from v_new")
        #
        # else:  ############# 입력값 하나라도 있을 경우 조건 검색
        #
        #
        #     self.tableWidget.setRowCount(len(allrows))  # 테이블 행 갯수 지정
        #     self.tableWidget.setColumnCount(11)  # 테이블 열 갯수 지정
        #     self.tableWidget.setHorizontalHeaderLabels(
        #         ['모델명', '일련번호', '상태', '일자', '지급내용', '인수자', '반납내용', '반납자', '확인자', '비고'])  # 헤드이름
        #     ####검색 데이터 표출
        #
        #     row_cnt = 0
        #
        #     for row in allrows:
        #         if row is None:
        #             break
        #         col_cnt = 0
        #         for col in row:
        #             if col == "None":  # 일자 null값 공백으로 표출
        #                 self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(""))
        #             else:
        #                 self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(col))
        #             col_cnt += 1
        #         row_cnt += 1
        #     self.tableWidget.setColumnHidden(10, True)  ####마지막 기록 순번 열 숨김

    def dialog_open(self):  # 수정 창 오픈

        print("수정클릭")
        try:

            standard = [selected_data[x].strip() for x in range(11)]
            # print(standard)
            ####초기 선택값 있는 경우 실행, selected_data는 수정창 테이블 한 번 표출 후 공백으로 초기화
            if any(standard):

                self.dialog = Ui_Update()  ## 수정창 클래스 생성

                self.close()  ###메인윈도우 닫음

                self.dialog.exec()  ###수정창 지속
                self.show()  ### 수정창 닫으면 메인윈도우 다시 표출

                self.textBrowser.setText(
                    self.db.execute_one("SELECT COUNT(*) from v_new WHERE state = '지급';")[0])  ##### 개수 표출
                ##### 미지급수량 세기
                self.textBrowser_2.setText(
                    self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('반납','none');")[0])  ##### 개수 표출
                ##### 총수량 세기
                self.textBrowser_3.setText(self.db.execute_all("select * from v_new"))  ##### 개수 표출

                self.tableWidget.clear()
                text_serial = self.textEdit.toPlainText()  # 사용자 일련번호 입력값 받아오기
                text_model = self.combo_change()  # 모델명
                text_recieve = self.textEdit_4.toPlainText()  # 인수자
                text_state = self.combo2_change()  # 지급여부
                text_date = self.textEdit_6.toPlainText()
                text_return = self.textEdit_7.toPlainText()
                text_list = [text_serial.strip(), text_model.strip(), text_recieve.strip(), text_state.strip(),
                             text_date.strip(), text_return.strip()]  # strip함수 : 공백->0 리턴

                if not any(text_list):  # 모든 칸에 입력 없을 경우 전체 헤드셋 표출
                    self.showTableListAll("select * from v_new")

                else:  ############# 입력값 하나라도 있을 경우 조건 검색
                    self.tableWidget.clear()  # 테이블 위젯 초기화
                    ####입력이 공백일 경우 해당 조건 패스
                    if not text_serial.strip():
                        text_serial = 'serial_num'
                    else:
                        text_serial = ("'{}'".format(text_serial))
                    if not text_model.strip():
                        text_model = 'model_name'
                    else:
                        text_model = ("'{}'".format(text_model))
                    if not text_recieve.strip():
                        text_recieve = 'recieve_person'
                    else:
                        text_recieve = ("'{}'".format(text_recieve))
                    if not text_state.strip():
                        text_state = 'state'
                    else:
                        text_state = ("'{}'".format(text_state))
                    if not text_date.strip():
                        text_date = 'state_date'
                    else:
                        try:
                            print(datetime.datetime.strptime(text_date, "%Y-%m-%d"))
                            text_date = ("'{}'".format(text_date))
                        except ValueError:
                            QMessageBox.warning(self, '경고', '날짜 형식은 YYYY-MM-DD로 입력해주세요.')
                            text_date = 'state_date'
                    if not text_return.strip():
                        text_return = 'return_person'
                    else:
                        text_return = ("'{}'".format(text_return))

                    sql = "SELECT * FROM v_new WHERE serial_num = {} AND model_name = {} \
                                       AND recieve_person = {} AND state = {} \
                                       AND state_date = {} AND return_person = {} \
                                       ORDER BY state_date DESC".format(text_serial, text_model, text_recieve,
                                                                        text_state, text_date,
                                                                        text_return)
                    ####입력이 공백일 경우 해당 조건 패스
                    self.showTableListAll(sql)

            elif not any(standard):  #####selected_data 값 없는 경우

                QMessageBox.warning(self, '경고', '수정할 데이터를 선택하세요.')


        except NameError:  ####selected_data 값 없는 경우
            QMessageBox.warning(self, '경고', '수정할 데이터를 선택하세요.')

    def showlist(self):  #### '검색'버튼 클릭 showlist함수
        # self.tableWidget.setStyleSheet("border-style:solid;"
        #                               "border-width: 2px;"
        #                               "border-color: #9DCFFF;"
        #                               "border-radius: 3p")
        ##### 지급수량 세기
        self.textBrowser.setText(
            str(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('지급');")[0]))  ##### 개수 표출
        ##### 미지급수량 세기
        self.textBrowser_2.setText(
            str(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('반납','none');")[0]))  ##### 개수 표출
        ##### 총수량 세기
        rowcount = self.db.execute_all("select * from v_new")
        print(rowcount)
        self.textBrowser_3.setText(str(len(rowcount)))  ##### 개수 표출

        global selected_data
        print("검색버튼 누름")

        # 등재된 전체 헤드셋 정보 표출

        text_serial = self.textEdit.toPlainText()  # 사용자 일련번호 입력값 받아오기
        text_model = self.combo_change()  # 모델명
        text_recieve = self.textEdit_4.toPlainText()  # 인수자
        text_state = self.combo2_change()  # 지급여부
        text_date = self.textEdit_6.toPlainText()
        text_return = self.textEdit_7.toPlainText()
        text_list = [text_serial.strip(), text_model.strip(), text_recieve.strip(), text_state.strip(),
                     text_date.strip(), text_return.strip()]

        if not any(text_list):  ########## 모든 칸에 입력 없을 경우 전체 헤드셋 표출
            print("모든 조건 입력없음")
            allrows = self.db.execute_all("select * from v_new")
            self.tableWidget.setRowCount(len(allrows))  # 테이블 행 갯수 지정
            self.tableWidget.setColumnCount(11)  # 테이블 열 갯수 지정
            self.tableWidget.setHorizontalHeaderLabels(
                ['모델명', '일련번호', '상태', '일자', '지급내용', '인수자', '반납내용', '반납자', '확인자', '비고'])  # 위젯 헤드이름

            row_cnt = 0
            print(allrows)
            for row in allrows:
                if row is None:
                    break
                col_cnt = 0
                for col in row:
                    if col == "None":  # 일자 null값 공백으로 표출
                        self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(""))
                    else:
                        self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(col))
                    col_cnt += 1
                row_cnt += 1

            self.tableWidget.setColumnHidden(10, True)  ####마지막 기록 순번 열 숨김

            self.tableWidget.resizeColumnToContents(4)  # 4번째 열 너비 조정
            self.tableWidget.resizeColumnToContents(6)  # 6번째 열 너비 조정
            self.tableWidget.resizeColumnToContents(9)  # 9번째 열 너비 조정

            selected_data = ['', '', '', '', '', '', '', '', '', '', '']
            print("발생:", selected_data)


        else:  ############# 입력값 하나라도 있을 경우
            self.tableWidget.clear()  # 테이블 위젯 초기화
            ####입력이 공백일 경우 해당 조건 패스
            if not text_serial.strip():
                text_serial = 'serial_num'
            else:
                text_serial = ("'{}'".format(text_serial))
            if not text_model.strip():
                text_model = 'model_name'
            else:
                text_model = ("'{}'".format(text_model))
            if not text_recieve.strip():
                text_recieve = 'recieve_person'
            else:
                text_recieve = ("'{}'".format(text_recieve))
            if not text_state.strip():
                text_state = 'state'
            else:
                text_state = ("'{}'".format(text_state))
            if not text_date.strip():
                text_date = 'state_date'
            else:
                try:
                    print(datetime.datetime.strptime(text_date, "%Y-%m-%d"))
                    text_date = ("'{}'".format(text_date))
                except ValueError:
                    QMessageBox.warning(self, '경고', '날짜 형식은 YYYY-MM-DD로 입력해주세요.')
                    text_date = 'state_date'
            if not text_return.strip():
                text_return = 'return_person'
            else:
                text_return = ("'{}'".format(text_return))

            sql = "SELECT * FROM v_new WHERE serial_num = {} AND model_name = {} \
                   AND recieve_person = {} AND state = {} \
                   AND state_date = {} AND return_person = {} \
                   ORDER BY state_date DESC".format(text_serial, text_model, text_recieve, text_state, text_date,
                                                    text_return)

            print(sql)
            ####테이블 행열 지정
            self.showTableListAll(sql)

            selected_data = ['', '', '', '', '', '', '', '', '', '', '']
            print("발생:", selected_data)

    def row_click(self, logicalIndex):  ###사용자 행 헤더 클릭시

        global row_index, selected_data

        row_index = logicalIndex + 1

        print(str(row_index) + '번째 행')
        try:
            # 수정창 클래스에 전달
            data1 = self.tableWidget.item(row_index - 1, 0).text()
            data2 = self.tableWidget.item(row_index - 1, 1).text()
            data3 = self.tableWidget.item(row_index - 1, 2).text()
            data4 = self.tableWidget.item(row_index - 1, 3).text()
            data5 = self.tableWidget.item(row_index - 1, 4).text()
            data6 = self.tableWidget.item(row_index - 1, 5).text()
            data7 = self.tableWidget.item(row_index - 1, 6).text()
            data8 = self.tableWidget.item(row_index - 1, 7).text()
            data9 = self.tableWidget.item(row_index - 1, 8).text()
            data10 = self.tableWidget.item(row_index - 1, 9).text()
            data11 = self.tableWidget.item(row_index - 1, 10).text()  # 기록 인덱스 값

            print(data11)
            selected_data = [data1, data2, data3, data4, data5, data6, data7, data8, data9, data10, data11]

            return row_index, selected_data
        except AttributeError:
            print("AttributeError")

    def cellclicked_event(self, row, col):
        global selected_data, row_index
        selected_data = ['', '', '', '', '', '', '', '', '', '', '']
        print("발생:", selected_data)

    def headerclicked_event(self, logicalIndex):
        global selected_data, row_index
        selected_data = ['', '', '', '', '', '', '', '', '', '', '']
        print("발생:", selected_data)

    def delete(self):  #### '삭제'버튼 클릭 이벤트
        global selected_data
        standard = selected_data[1].strip()
        if standard:  ##일련번호 공백이 아닐경우

            answer = self.Question_event()  ### yes or no
            if answer:

                serial_data = selected_data[1]  ##선택 행 일련번호 값 가져오
                self.db.excute(f"UPDATE headset SET valid = 0 WHERE serial_num = '{serial_data}';")
                QMessageBox.about(self, '알림', '일련번호 : {} 헤드셋이 삭제되었습니다.'.format(serial_data))

                ########################    삭제 후 메인창 다시 검색 ###################################
                ##### 지급수량 세기
                self.textBrowser.setText(
                    str(self.db.execute_one("SELECT COUNT(*) from v_new WHERE state = '지급';")[0]))  ##### 개수 표출
                ##### 미지급수량 세기
                self.textBrowser_2.setText(str(
                    self.db.execute_one("SELECT COUNT(*) from v_new WHERE state IN ('반납','none');")[0]))  ##### 개수 표출
                ##### 총수량 세기
                self.textBrowser_3.setText(str(len(self.db.execute_all("select * from v_new"))))  ##### 개수 표출

                self.tableWidget.clear()
                text_serial = self.textEdit.toPlainText()  # 사용자 일련번호 입력값 받아오기
                text_model = self.combo_change()  # 모델명
                text_recieve = self.textEdit_4.toPlainText()  # 인수자
                text_state = self.combo2_change()  # 지급여부
                text_date = self.textEdit_6.toPlainText()
                text_return = self.textEdit_7.toPlainText()
                text_list = [text_serial.strip(), text_model.strip(), text_recieve.strip(), text_state.strip(),
                             text_date.strip(), text_return.strip()]

                # self.tableWidget.setStyleSheet("border-style:solid;"
                #                   "border-width: 2px;"
                #                   "border-color: #9DCFFF;"
                #                   "border-radius: 3p")

                if not any(text_list):  ########## 모든 칸에 입력 없을 경우 전체 헤드셋 표출
                    allrows = self.db.execute_all("select * from v_new")
                    self.tableWidget.setRowCount(len(allrows))  # 테이블 행 갯수 지정
                    self.tableWidget.setColumnCount(11)  # 테이블 열 갯수 지정
                    self.tableWidget.setHorizontalHeaderLabels(
                        ['모델명', '일련번호', '상태', '일자', '지급내용', '인수자', '반납내용', '반납자', '확인자', '비고'])  # 위젯 헤드이름
                    # 등재된 전체 헤드셋 정보 표출
                    row_cnt = 0
                    for row in allrows:
                        if row is None:
                            break
                        col_cnt = 0
                        for col in row:
                            if col == "None":  # 일자 null값 공백으로 표출
                                self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(""))
                            else:
                                self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(col))
                            col_cnt += 1
                        row_cnt += 1
                    self.tableWidget.setColumnHidden(10, True)  ####마지막 기록 순번 열 숨김
                else:  ############# 입력값 하나라도 있을 경우 조건 검
                    self.tableWidget.clear()  # 테이블 위젯 초기화
                    ####입력이 공백일 경우 해당 조건 패스
                    if not text_serial.strip():
                        text_serial = 'serial_num'
                    else:
                        text_serial = ("'{}'".format(text_serial))
                    if not text_model.strip():
                        text_model = 'model_name'
                    else:
                        text_model = ("'{}'".format(text_model))
                    if not text_recieve.strip():
                        text_recieve = 'recieve_person'
                    else:
                        text_recieve = ("'{}'".format(text_recieve))
                    if not text_state.strip():
                        text_state = 'state'
                    else:
                        text_state = ("'{}'".format(text_state))
                    if not text_date.strip():
                        text_date = 'state_date'
                    else:
                        try:
                            print(datetime.datetime.strptime(text_date, "%Y-%m-%d"))
                            text_date = ("'{}'".format(text_date))
                        except ValueError:
                            QMessageBox.warning(self, '경고', '날짜 형식은 YYYY-MM-DD로 입력해주세요.')
                            text_date = 'state_date'
                    if not text_return.strip():
                        text_return = 'return_person'
                    else:
                        text_return = ("'{}'".format(text_return))

                    sql = "SELECT * FROM v_new WHERE serial_num = {} AND model_name = {} \
                               AND recieve_person = {} AND state = {} \
                               AND state_date = {} AND return_person = {} \
                               ORDER BY state_date DESC".format(text_serial, text_model, text_recieve, text_state,
                                                                text_date, text_return)
                    ####테이블 행열 지정
                    allrows = self.db.execute_all(sql)
                    self.tableWidget.setRowCount(len(allrows))  # 테이블 행 갯수 지정
                    self.tableWidget.setColumnCount(11)  # 테이블 열 갯수 지정
                    self.tableWidget.setHorizontalHeaderLabels(
                        ['모델명', '일련번호', '상태', '일자', '지급내용', '인수자', '반납내용', '반납자', '확인자', '비고'])  # 헤드이
                    ####검색 데이터 표출
                    row_cnt = 0
                    for row in allrows:
                        if row is None:
                            break
                        col_cnt = 0
                        for col in row:
                            if col == "None":  # 일자 null값 공백으로 표출
                                self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(""))
                            else:
                                self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(col))
                            col_cnt += 1
                        row_cnt += 1
                    self.tableWidget.setColumnHidden(10, True)  ####마지막 기록 순번 열 숨김
            elif not answer:
                pass

        elif not standard:  ####선택값 공백인 경우
            QMessageBox.warning(self, '경고', '삭제할 데이터를 선택하세요.')

        selected_data = ['', '', '', '', '', '', '', '', '', '', '']

    def Question_event(self):

        serial_data = selected_data[1]

        ment = '일련번호 : {} 헤드셋을 삭제하시겠습니까?'.format(selected_data[1])
        buttonReply = QMessageBox.question(self, '알림', ment, QMessageBox.Yes | QMessageBox.No)

        if buttonReply == QMessageBox.Yes:
            print('Yes clicked.')
            return True
        elif buttonReply == QMessageBox.No:
            print('No clicked.')
            return False

    def export(self):

        try:
            filename, filter = QtWidgets.QFileDialog.getSaveFileName(self, 'Save file', 'headset.xlsx',
                                                                     "Excel files(*.xlsx)")
            print(str(filename))

            wb = openpyxl.Workbook()

            sheet = wb.active
            sheet.title = '헤드셋 리스트'

            # 헤더 '모델명','일련번호','상태','일자', '지급내용','인수자','반납내용','반납자','확인자','비고'
            sheet.cell(1, 1, '모델명')
            sheet.cell(1, 2, '일련번호')
            sheet.cell(1, 3, '상태')
            sheet.cell(1, 4, '일자')
            sheet.cell(1, 5, '지급내용')
            sheet.cell(1, 6, '인수자')
            sheet.cell(1, 7, '반납내용')
            sheet.cell(1, 8, '반납자')
            sheet.cell(1, 9, '확인자')
            sheet.cell(1, 10, '비고')

            # sheet.cell(1,1,'s')
            row = 0
            column = 0
            rowcount = self.tableWidget.rowCount()
            columncount = self.tableWidget.columnCount()
            print("행{},열{}".format(str(rowcount), str(columncount)))

            for row in range(0, rowcount):
                for column in range(0, columncount - 1):  ###10열 숨김

                    text = str(self.tableWidget.item(row, column).text())
                    sheet.cell(row + 2, column + 1, text)  ##엑셀 1,1 부터 인덱스 시작
                    column = + 1
                row = +1

            wb.save(filename)
            wb.close()
        except AttributeError:
            QMessageBox.warning(self, '경고', '다운로드할 데이터를 검색하세요')
        except FileNotFoundError:
            print("FileNotFoundError")
            # except Exception as e:
            #    print("Error Writing file.")

    # except Exception as e:
    #    print("Error Saving file.")

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "헤드셋 관리 프로그램"))  # 윈도우창 이름
        self.label.setText(_translate("MainWindow", "검색 리스트"))
        self.groupBox.setTitle(_translate("MainWindow", "관리현황"))
        self.label_2.setText(_translate("MainWindow", "지급 수량 "))
        self.label_4.setText(_translate("MainWindow", "미지급 수량 "))
        self.label_5.setText(_translate("MainWindow", "총수량 "))
        self.groupBox_2.setTitle(_translate("MainWindow", "입력"))
        self.label_6.setText(_translate("MainWindow", "일련번호"))
        self.label_7.setText(_translate("MainWindow", "모델명"))
        self.label_8.setText(_translate("MainWindow", "인수자"))
        self.label_9.setText(_translate("MainWindow", "지급여부"))
        self.label_10.setText(_translate("MainWindow", "일자"))
        self.label_11.setText(_translate("MainWindow", "반납자"))
        self.pushButton.setText(_translate("MainWindow", "검색"))
        self.pushButton_2.setText(_translate("MainWindow", "수정"))
        self.pushButton_3.setText(_translate("MainWindow", "삭제"))
        self.pushButton_4.setText(_translate("MainWindow", "신규등록"))
        self.pushButton_5.setText(_translate("MainWindow", "엑셀\n다운로드"))

    def getConditions(self):
        text_serial = self.textEdit.toPlainText()  # 사용자 일련번호 입력값 받아오기
        text_model = self.combo_change()  # 모델명
        text_recieve = self.textEdit_4.toPlainText()  # 인수자
        text_state = self.combo2_change()  # 지급여부
        text_date = self.textEdit_6.toPlainText()
        text_return = self.textEdit_7.toPlainText()
        text_list = [text_serial.strip(), text_model.strip(), text_recieve.strip(), text_state.strip(),
                     text_date.strip(), text_return.strip()]  # strip함수 : 공백->0 리턴
        return text_list

    def showTableListAll(self, query):
        allrows = self.db.execute_all(query)
        self.tableWidget.setRowCount(len(allrows))  # 테이블 행 갯수 지정
        self.tableWidget.setColumnCount(11)  # 테이블 열 갯수 지정
        self.tableWidget.setHorizontalHeaderLabels(
            ['모델명', '일련번호', '상태', '일자', '지급내용', '인수자', '반납내용', '반납자', '확인자', '비고'])  # 위젯 헤드이름
        # 등재된 전체 헤드셋 정보 표출
        # cur.execute("select * from v_new")
        row_cnt = 0
        for row in allrows:
            if row is None:
                break
            col_cnt = 0
            for col in row:
                if col == "None":  # 일자 null값 공백으로 표출
                    self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(""))
                else:
                    self.tableWidget.setItem(row_cnt, col_cnt, QTableWidgetItem(col))
                col_cnt += 1
            row_cnt += 1

        self.tableWidget.setColumnHidden(10, True)

    # def conditinalSearch(self, text_list):
    #     if not text_serial.strip():
    #         text_serial = 'serial_num'
    #     else:
    #         text_serial = ("'{}'".format(text_serial))
    #     if not text_model.strip():
    #         text_model = 'model_name'
    #     else:
    #         text_model = ("'{}'".format(text_model))
    #     if not text_recieve.strip():
    #         text_recieve = 'recieve_person'
    #     else:
    #         text_recieve = ("'{}'".format(text_recieve))
    #     if not text_state.strip():
    #         text_state = 'state'
    #     else:
    #         text_state = ("'{}'".format(text_state))
    #     if not text_date.strip():
    #         text_date = 'state_date'
    #     else:
    #         try:
    #             print(datetime.datetime.strptime(text_date, "%Y-%m-%d"))
    #             text_date = ("'{}'".format(text_date))
    #         except ValueError:
    #             QMessageBox.warning(self, '경고', '날짜 형식은 YYYY-MM-DD로 입력해주세요.')
    #             text_date = 'state_date'
    #     if not text_return.strip():
    #         text_return = 'return_person'
    #     else:
    #         text_return = ("'{}'".format(text_return))
    #
    #     sql = "SELECT * FROM v_new WHERE serial_num = {} AND model_name = {} \
    #     AND recieve_person = {} AND state = {} \
    #     AND state_date = {} AND return_person = {} \
    #     ORDER BY state_date DESC".format(text_serial, text_model, text_recieve, text_state, text_date,
    #                                      text_return)
    #     allrows = self.db.execute_all(sql)

##### qtdesigner 자동생성x
class kinwriter(QMainWindow, Ui_MainWindow, QTableWidgetItem):

    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.showlist()
        self.show()
        # self.timer = QTimer(self)
        # self.timer.setSingleShot(False)
        # self.timer.setInterval(5000) # in milliseconds, so 5000 = 5 seconds
        # # self.timer.timeout.connect(self.start_Macro)
        # self.timer.start()i0nscn2kdlr2k

        # print(self.hasMouseTracking())


app = QApplication([sys.argv])
# ex = Ui_MainWindow()

sn = kinwriter()  # 인스턴스 생성
QApplication.processEvents()
sys.exit(app.exec_())
